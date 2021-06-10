import openpyxl


listdata = [[],[]]

fn = '../src/微波爐財編.xlsx'
wb = openpyxl.load_workbook(fn)
wb.active = 0

ws = wb.active

for row in ws['G2':'G54']:
    for cell in row:
        #print(cell.value,end=' ')
        listdata[0].append(cell.value)


for row in ws['I2':'I54']:
    for cell in row:
        #print(cell.value,end=' ')
        if (cell.value == None):
            listdata[1].append(" ")
        else:
            listdata[1].append(cell.value)


for index in range(len(listdata[0])):
    print(listdata[0][index] , " " , listdata[1][index])


