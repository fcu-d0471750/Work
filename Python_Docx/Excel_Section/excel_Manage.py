# 導入 openpyxl
import openpyxl
# 導入doc管理
from Docx_Section import docx_Manage as DM
# 導入資料庫管理
from Store_DB.DB_Manage import use_store_db

fn = '../src/110-06-微波爐財編-1.xlsx'
wb = openpyxl.load_workbook(fn)
wb.active = 0

ws = wb.active

listdata = [[],[]]
USE_DB = []

# 執行
def do_excel(require_area,filename,require_date,require_store,*name_number):
    # 讀取 Excel
    listdata = read_excel(*name_number)
    # 擴展成ㄧ般格式
    USE_DB = use_store_db(listdata,require_area)
    # 執行 Doc
    DM.do_doc(require_area,filename,require_date,require_store,USE_DB)

# 讀取Excel
def read_excel(*start_end):
    Temp = [[], []]
    for i in range(len(start_end)):

        # 讀取店家名稱
        for row in ws[start_end[i][0]:start_end[i][1]]:
            for cell in row:
                # 空值
                if(cell.value==None):Temp[0].append("xxx")
                else : Temp[0].append(cell.value)

        # 讀取財產編號
        for row in ws[start_end[i][2]:start_end[i][3]]:
            for cell in row:
                # 空值
                if (cell.value == None):
                    Temp[1].append("xxx")
                else:
                    Temp[1].append(cell.value)

    return Temp
